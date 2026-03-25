from flask import Flask, render_template, request, jsonify, redirect, url_for, send_file
import json
import os
import uuid
from datetime import date
import io

from scheduler import generate_schedule, get_period_dates, calculate_summary

app = Flask(__name__)

DATA_DIR      = 'data'
EMPLOYEES_FILE = os.path.join(DATA_DIR, 'employees.json')
SETTINGS_FILE  = os.path.join(DATA_DIR, 'settings.json')
SCHEDULES_DIR  = os.path.join(DATA_DIR, 'schedules')

MONTH_NAMES = [
    '', 'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
    'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember'
]


# ─── Data helpers ─────────────────────────────────────────────────────────────

def init_data():
    os.makedirs(DATA_DIR, exist_ok=True)
    os.makedirs(SCHEDULES_DIR, exist_ok=True)
    if not os.path.exists(EMPLOYEES_FILE):
        _write(EMPLOYEES_FILE, [])
    if not os.path.exists(SETTINGS_FILE):
        _write(SETTINGS_FILE, {
            'branches': [
                {
                    'id':       str(uuid.uuid4()),
                    'name':     'Cabang Utama',
                    'jobdesks': ['Kasir', 'Security', 'Admin'],
                }
            ],
            'shift_pagi':  {'start': '07:00', 'end': '15:00'},
            'shift_siang': {'start': '15:00', 'end': '23:00'},
        })


import threading

db_lock = threading.Lock()

def _read(path):
    with db_lock:
        with open(path, encoding='utf-8') as f:
            return json.load(f)

def _write(path, data):
    with db_lock:
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)


def load_employees():
    return _read(EMPLOYEES_FILE)


def save_employees(d):
    _write(EMPLOYEES_FILE, d)


def load_employees_for_period(period):
    """Return daftar pegawai yang relevan untuk sebuah period (filter per cabang)."""
    emps = load_employees()
    bid  = period.get('branch_id', '') if period else ''
    return [e for e in emps if e.get('branch') == bid] if bid else emps

def load_settings():
    s = _read(SETTINGS_FILE)
    # ── Backward compat: settings lama tanpa 'branches' ──────────────────────
    if 'branches' not in s:
        s['branches'] = [{
            'id':       str(uuid.uuid4()),
            'name':     'Cabang Utama',
            'jobdesks': s.get('jobdesks', ['Kasir', 'Security', 'Admin']),
        }]
        _write(SETTINGS_FILE, s)
    return s


def save_settings(d):
    _write(SETTINGS_FILE, d)


def schedule_path(pk):
    return os.path.join(SCHEDULES_DIR, f'{pk}.json')


def load_schedule(pk):
    p = schedule_path(pk)
    return _read(p) if os.path.exists(p) else None


def save_schedule(pk, data):
    _write(schedule_path(pk), data)


def list_schedules():
    if not os.path.exists(SCHEDULES_DIR):
        return []
    return sorted(
        [f[:-5] for f in os.listdir(SCHEDULES_DIR) if f.endswith('.json')],
        reverse=True,
    )


# ─── Branch helpers ───────────────────────────────────────────────────────────

def get_branch(settings, bid):
    """Cari branch by id, return None jika tidak ada."""
    for b in settings.get('branches', []):
        if b['id'] == bid:
            return b
    return None


def branch_map(settings):
    """Return {branch_id: branch_dict}."""
    return {b['id']: b for b in settings.get('branches', [])}


def group_by_branch(emps, settings):
    """
    Return {branch_id: {'branch': branch_dict, 'by_jobdesk': {jd: [emp, ...]}}}.
    Pegawai tanpa branch yang dikenal dimasukkan ke key ''.
    """
    bmap   = branch_map(settings)
    result = {}

    # Buat slot untuk setiap branch yang ada di settings
    for b in settings.get('branches', []):
        result[b['id']] = {'branch': b, 'by_jobdesk': {}}

    for emp in emps:
        bid = emp.get('branch', '')
        jd  = emp['jobdesk']
        if bid not in result:
            # Branch tidak dikenal (data lama) → slot khusus
            result[bid] = {
                'branch':    {'id': bid, 'name': 'Lainnya', 'jobdesks': []},
                'by_jobdesk': {},
            }
        result[bid]['by_jobdesk'].setdefault(jd, []).append(emp)

    return result


def all_jobdesks(settings):
    """Kumpulkan semua jobdesk dari semua branch (unik)."""
    seen, out = set(), []
    for b in settings.get('branches', []):
        for jd in b.get('jobdesks', []):
            if jd not in seen:
                seen.add(jd)
                out.append(jd)
    return out


# ─── Routes: General ──────────────────────────────────────────────────────────

@app.route('/')
def index():
    settings  = load_settings()
    bmap      = branch_map(settings)
    schedules = []
    for pk in list_schedules():
        parts = pk.split('-')
        y, m  = int(parts[0]), int(parts[1])
        # period_key: YYYY-MM or YYYY-MM-BID
        bid   = parts[2] if len(parts) > 2 else ''
        bname = bmap.get(bid, {}).get('name', 'Semua Cabang') if bid else 'Semua Cabang'
        data  = load_schedule(pk)
        schedules.append({
            'key':         pk,
            'label':       f"{MONTH_NAMES[m]} {y}",
            'branch_name': bname,
            'generated':   data.get('generated', False) if data else False,
        })
    today = date.today()
    return render_template('index.html', schedules=schedules, today=today,
                           month_names=MONTH_NAMES[1:], settings=settings)

# ─── Routes: Employees ────────────────────────────────────────────────────────

@app.route('/employees')
def employees():
    emps     = load_employees()
    settings = load_settings()
    bmap     = branch_map(settings)

    # Tambahkan branch_name ke setiap employee (untuk tampilan)
    for e in emps:
        bid = e.get('branch', '')
        e['branch_name'] = bmap.get(bid, {}).get('name', 'Tidak diketahui')

    by_branch   = group_by_branch(emps, settings)
    jd_all      = all_jobdesks(settings)

    return render_template(
        'employees.html',
        employees=emps,
        jobdesks=jd_all,          # backward compat untuk template lama
        by_branch=by_branch,
        settings=settings,
    )


@app.route('/employees/add', methods=['POST'])
def add_employee():
    emps = load_employees()
    new_emp = {
        'id':      str(uuid.uuid4()),
        'name':    request.form['name'].strip(),
        'branch':  request.form.get('branch', ''),
        'jobdesk': request.form['jobdesk'],
        'gender':  request.form['gender'],
    }
    emps.append(new_emp)
    save_employees(emps)
    # JSON response for AJAX / CSR
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'success': True, 'employee': new_emp})
    return redirect(url_for('employees'))


@app.route('/employees/inline/<eid>', methods=['POST'])
def inline_edit_employee(eid):
    data = request.get_json()
    if not data:
        return jsonify({'success': False})
    
    field = data.get('field')
    value = data.get('value')
    
    emps = load_employees()
    for e in emps:
        if e['id'] == eid:
            if field in ['name', 'jobdesk', 'gender']:
                e[field] = value
            save_employees(emps)
            return jsonify({'success': True})
            
    return jsonify({'success': False})


@app.route('/employees/edit/<eid>', methods=['POST'])
def edit_employee(eid):
    emps = load_employees()
    for e in emps:
        if e['id'] == eid:
            e['name']    = request.form['name'].strip()
            e['branch']  = request.form.get('branch', e.get('branch', ''))
            e['jobdesk'] = request.form['jobdesk']
            e['gender']  = request.form['gender']
            break
    save_employees(emps)
    return redirect(url_for('employees'))


@app.route('/employees/delete/<eid>', methods=['POST'])
def delete_employee(eid):
    emps = [e for e in load_employees() if e['id'] != eid]
    save_employees(emps)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'success': True})
    return redirect(url_for('employees'))


# ─── Routes: Settings ─────────────────────────────────────────────────────────

@app.route('/settings', methods=['GET', 'POST'])
def settings():
    if request.method == 'POST':
        s = load_settings()
        s['shift_pagi']  = {
            'start': request.form['pagi_start'],
            'end':   request.form['pagi_end'],
        }
        s['shift_siang'] = {
            'start': request.form['siang_start'],
            'end':   request.form['siang_end'],
        }
        save_settings(s)
        return redirect(url_for('settings'))
    return render_template('settings.html', settings=load_settings())


@app.route('/settings/branch/add', methods=['POST'])
def add_branch():
    s = load_settings()
    s.setdefault('branches', []).append({
        'id':       str(uuid.uuid4()),
        'name':     request.form['name'].strip(),
        'jobdesks': [],
    })
    save_settings(s)
    return redirect(url_for('settings'))


@app.route('/settings/branch/<bid>/edit', methods=['POST'])
def edit_branch(bid):
    s = load_settings()
    b = get_branch(s, bid)
    if b:
        b['name'] = request.form['name'].strip()
        raw = request.form.get('jobdesks', '')
        b['jobdesks'] = [j.strip() for j in raw.split('\n') if j.strip()]
        save_settings(s)
    return redirect(url_for('settings'))


@app.route('/settings/branch/<bid>/delete', methods=['POST'])
def delete_branch(bid):
    s = load_settings()
    s['branches'] = [b for b in s.get('branches', []) if b['id'] != bid]
    save_settings(s)
    return redirect(url_for('settings'))


# JSON API — untuk dropdown jobdesk dinamis saat input pegawai
@app.route('/api/branch/<bid>/jobdesks')
def api_branch_jobdesks(bid):
    s = load_settings()
    b = get_branch(s, bid)
    return jsonify({'jobdesks': b['jobdesks'] if b else []})


# ─── Routes: Schedule ─────────────────────────────────────────────────────────

@app.route('/schedule/new', methods=['POST'])
def new_schedule():
    year   = int(request.form['year'])
    month  = int(request.form['month'])
    bid    = request.form.get('branch', '')
    # period_key bersifat unik per cabang
    pk = f'{year}-{month:02d}-{bid}' if bid else f'{year}-{month:02d}'

    period = load_schedule(pk)
    if not period:
        settings = load_settings()
        emps_all = load_employees()
        # Filter pegawai sesuai cabang yang dipilih
        emps = [e for e in emps_all if e.get('branch') == bid] if bid else emps_all
        dates = get_period_dates(year, month)
        period = {
            'period_key': pk,
            'year':       year,
            'month':      month,
            'branch_id':  bid,
            'label':      f"{MONTH_NAMES[month]} {year}",
            'dates':      dates,
            'off_days':   {e['id']: [] for e in emps},
            'cuti_days':  {e['id']: [] for e in emps},
            'schedule':   {e['id']: {} for e in emps},
            'generated':  False,
        }
        save_schedule(pk, period)

    if period.get('generated'):
        return redirect(url_for('schedule_view', pk=pk))
    return redirect(url_for('schedule_setup', pk=pk))


@app.route('/schedule/<pk>/setup')
def schedule_setup(pk):
    period = load_schedule(pk)
    if not period:
        return redirect(url_for('index'))

    if period.get('generated'):
        return redirect(url_for('schedule_view', pk=pk))

    emps      = load_employees_for_period(period)
    settings  = load_settings()
    by_branch = group_by_branch(emps, settings)

    summary = calculate_summary(period, emps)
    return render_template('schedule_setup.html', period=period, by_branch=by_branch, summary=summary)


@app.route('/schedule/<pk>/save_off', methods=['POST'])
def save_off(pk):
    # ... (content remains the same, skipped for brevity to not break existing lines, let's include it all to be safe)
    period = load_schedule(pk)
    if not period:
        return jsonify({'success': False})
        
    eid = request.form.get('eid')
    if not eid:
        return jsonify({'success': False, 'error': 'No EID provided'})

    if 'off_days' not in period: period['off_days'] = {}
    if 'cuti_days' not in period: period['cuti_days'] = {}
    
    period['off_days'][eid]  = request.form.getlist('off[]')
    period['cuti_days'][eid] = request.form.getlist('cuti[]')
    
    # Sync visual ke schedule data jika ada, supaya generator & UI tidak bingung
    if 'schedule' not in period: period['schedule'] = {}
    if eid not in period['schedule']: period['schedule'][eid] = {}
    
    # Hapus OFF/CUTI lama dari dict schedule untuk eid ini
    for d in list(period['schedule'][eid].keys()):
        if period['schedule'][eid][d] in ('OFF', 'CUTI'):
            period['schedule'][eid].pop(d, None)
            
    # Masukkan yang baru
    for d in period['off_days'][eid]:
        period['schedule'][eid][d] = 'OFF'
    for d in period['cuti_days'][eid]:
        period['schedule'][eid][d] = 'CUTI'

    save_schedule(pk, period)
    return jsonify({'success': True})


@app.route('/schedule/<pk>/auto_off_template', methods=['POST'])
def auto_off_template(pk):
    period = load_schedule(pk)
    if not period:
        return jsonify({'success': False})
        
    emps = load_employees_for_period(period)
    settings = load_settings()
    by_branch = group_by_branch(emps, settings)
    
    if 'off_days' not in period: period['off_days'] = {e['id']: [] for e in emps}
    if 'schedule' not in period: period['schedule'] = {e['id']: {} for e in emps}
    
    # Reset semua OFF yang pernah ada
    for e in emps:
        period['off_days'][e['id']] = []
        for d in list(period['schedule'].get(e['id'], {}).keys()):
            if period['schedule'][e['id']][d] == 'OFF':
                period['schedule'][e['id']].pop(d, None)
                
    dates = period['dates']
    
    for b_id, b_data in by_branch.items():
        for jd, jd_emps in b_data['by_jobdesk'].items():
            jd_emps = sorted(jd_emps, key=lambda x: x['name'])
            for i, emp in enumerate(jd_emps):
                eid = emp['id']
                # Spacing them out: Assign 1 OFF per 7 days, staggering the start day.
                # Batas maksimal 4 OFF per orang dalam satu periode.
                offset = i % 7
                off_count = 0
                for day_idx in range(offset, len(dates), 7):
                    if off_count >= 4:
                        break
                    d_str = dates[day_idx]
                    period['off_days'][eid].append(d_str)
                    period['schedule'][eid][d_str] = 'OFF'
                    off_count += 1

    save_schedule(pk, period)
    return jsonify({'success': True, 'reload': True, 'off_days': period['off_days']})


@app.route('/schedule/<pk>/generate', methods=['POST'])
def generate(pk):
    period = load_schedule(pk)
    if not period:
        return jsonify({'success': False, 'error': 'Period not found'})
    emps = load_employees_for_period(period)
    if not emps:
        return jsonify({'success': False, 'error': 'Tidak ada pegawai di cabang ini'})
    sched = generate_schedule(period, emps)
    period['schedule']  = sched
    period['generated'] = True
    save_schedule(pk, period)
    return jsonify({'success': True, 'redirect': url_for('schedule_view', pk=pk)})


@app.route('/schedule/<pk>')
def schedule_view(pk):
    period = load_schedule(pk)
    if not period:
        return redirect(url_for('index'))
    emps      = load_employees_for_period(period)
    settings  = load_settings()
    summary   = calculate_summary(period, emps)
    by_branch = group_by_branch(emps, settings)

    by_jd = {}
    for e in emps:
        by_jd.setdefault(e['jobdesk'], []).append(e)

    return render_template(
        'schedule_view.html',
        period=period, employees=emps,
        by_branch=by_branch,
        by_jobdesk=by_jd,
        settings=settings,
        summary=summary,
    )


@app.route('/schedule/<pk>/edit_cell', methods=['POST'])
def edit_cell(pk):
    data   = request.get_json()
    period = load_schedule(pk)
    if not period:
        return jsonify({'success': False})
        
    eid    = data['emp_id']
    d      = data['date']
    status = data['status']
    
    if status == '':
        period['schedule'].setdefault(eid, {}).pop(d, None)
    else:
        period['schedule'].setdefault(eid, {})[d] = status
        
    # Sync with off_days and cuti_days arrays
    off_list = period.setdefault('off_days', {}).setdefault(eid, [])
    cuti_list = period.setdefault('cuti_days', {}).setdefault(eid, [])
    
    if status == 'OFF':
        if d not in off_list: off_list.append(d)
        if d in cuti_list: cuti_list.remove(d)
    elif status == 'CUTI':
        if d not in cuti_list: cuti_list.append(d)
        if d in off_list: off_list.remove(d)
    else:
        if d in off_list: off_list.remove(d)
        if d in cuti_list: cuti_list.remove(d)
        
    save_schedule(pk, period)
    return jsonify({'success': True})


@app.route('/schedule/<pk>/delete', methods=['POST'])
def delete_schedule(pk):
    p = schedule_path(pk)
    if os.path.exists(p):
        os.remove(p)
    return redirect(url_for('index'))


# ─── Routes: Export ───────────────────────────────────────────────────────────

@app.route('/schedule/<pk>/export/excel')
def export_excel(pk):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    period   = load_schedule(pk)
    emps     = load_employees()
    settings = load_settings()
    dates    = period['dates']
    sched    = period.get('schedule', {})
    branches = settings.get('branches', [])
    
    bid_filter = request.args.get('bid')
    if bid_filter:
        branches = [b for b in branches if b['id'] == bid_filter]
        
    n_dates  = len(dates)

    # Kolom: Nama | Jobdesk | G | [dates] | PAGI | SIANG | OFF | CUTI
    DATE_START_COL = 4
    SUM_START_COL  = DATE_START_COL + n_dates
    TOTAL_COLS     = SUM_START_COL + 4

    COLORS = {
        'PAGI':    'FFD700',
        'SIANG':   '3A7BD5',
        'OFF':     'E74C3C',
        'CUTI':    '2ECC71',
        'branch':  '1A252F',
        'col_hdr': '2C3E50',
        'jd_hdr':  'D5E8F7',
        'alt':     'F7FAFD',
    }
    day_abbr = ['Sen', 'Sel', 'Rab', 'Kam', 'Jum', 'Sab', 'Min']

    thin        = Side(border_style='thin', color='CCCCCC')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Pre-group employees by branch → jobdesk ──────────────────────────────
    by_bjd = {}   # {branch_id: {jobdesk: [emp]}}
    for b in branches:
        by_bjd[b['id']] = {}
    for emp in emps:
        bid = emp.get('branch', '')
        jd  = emp['jobdesk']
        if bid not in by_bjd:
            by_bjd[bid] = {}
        by_bjd[bid].setdefault(jd, []).append(emp)

    # ── Style helpers ─────────────────────────────────────────────────────────
    def sc(cell, bg=None, bold=False, color='000000', size=9,
           h='center', v='center', wrap=False):
        """Style a cell."""
        if bg:
            cell.fill = PatternFill('solid', fgColor=bg)
        cell.font      = Font(bold=bold, color=color, size=size)
        cell.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
        cell.border    = thin_border
        return cell

    def write_col_headers(ws, row):
        sc(ws.cell(row, 1, 'Nama'),    bg=COLORS['col_hdr'], bold=True, color='FFFFFF')
        sc(ws.cell(row, 2, 'Jobdesk'), bg=COLORS['col_hdr'], bold=True, color='FFFFFF')
        sc(ws.cell(row, 3, 'G'),       bg=COLORS['col_hdr'], bold=True, color='FFFFFF')
        for ci, d in enumerate(dates, DATE_START_COL):
            do = date.fromisoformat(d)
            sc(ws.cell(row, ci, f"{do.strftime('%d/%m')}\n{day_abbr[do.weekday()]}"),
               bg=COLORS['col_hdr'], bold=True, color='FFFFFF', wrap=True)
        for i, lbl in enumerate(['PAGI', 'SIANG', 'OFF', 'CUTI']):
            sc(ws.cell(row, SUM_START_COL + i, lbl),
               bg=COLORS['col_hdr'], bold=True, color='FFFFFF')
        ws.row_dimensions[row].height = 34

    def write_emp_row(ws, row, emp, alt=False):
        eid = emp['id']
        bg  = COLORS['alt'] if alt else None
        sc(ws.cell(row, 1, emp['name']),    bg=bg, h='left',   size=9)
        sc(ws.cell(row, 2, emp['jobdesk']), bg=bg, h='left',   size=9)
        sc(ws.cell(row, 3, emp['gender']),  bg=bg,             size=9)
        counts = {'PAGI': 0, 'SIANG': 0, 'OFF': 0, 'CUTI': 0}
        for ci, d in enumerate(dates, DATE_START_COL):
            st = sched.get(eid, {}).get(d, '')
            c  = sc(ws.cell(row, ci, st or '-'), bg=bg, size=8)
            if st in COLORS:
                c.fill = PatternFill('solid', fgColor=COLORS[st])
                txt    = 'FFFFFF' if st in ('SIANG', 'OFF') else '333333'
                c.font = Font(size=8, bold=True, color=txt)
                c.border = thin_border
            if st in counts:
                counts[st] += 1
        for i, lbl in enumerate(['PAGI', 'SIANG', 'OFF', 'CUTI']):
            sc(ws.cell(row, SUM_START_COL + i, counts[lbl]), bg=bg, bold=True, size=9)
        ws.row_dimensions[row].height = 18

    def set_col_widths(ws):
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 13
        ws.column_dimensions['C'].width = 5
        for ci in range(DATE_START_COL, SUM_START_COL):
            ws.column_dimensions[get_column_letter(ci)].width = 6.5
        for ci in range(SUM_START_COL, SUM_START_COL + 4):
            ws.column_dimensions[get_column_letter(ci)].width = 7

    # ── Buat workbook ─────────────────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)   # hapus default sheet

    for branch in branches:
        bid    = branch['id']
        jd_map = by_bjd.get(bid, {})
        if not any(jd_map.values()):
            continue   # lewati cabang yang tidak ada pegawainya

        # Nama sheet: max 31 karakter, no special chars
        sheet_name = branch['name'][:31].replace('/', '-').replace('\\', '-')
        ws = wb.create_sheet(title=sheet_name)
        set_col_widths(ws)

        # Baris 1 — judul cabang
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1,   end_column=TOTAL_COLS - 1)
        title_text = (f"JADWAL SHIFT  ·  {branch['name'].upper()}"
                      f"  ·  {period.get('label', pk).upper()}")
        sc(ws.cell(1, 1, title_text),
           bg=COLORS['branch'], bold=True, color='FFFFFF', size=13, wrap=False)
        ws.row_dimensions[1].height = 28

        # Baris 2 — header kolom
        write_col_headers(ws, 2)

        current_row = 3

        # Urutan jobdesk sesuai setting cabang
        jd_order = [jd for jd in branch.get('jobdesks', []) if jd in jd_map]
        # Tambahkan jobdesk yang ada tapi tidak di setting (jaga-jaga)
        for jd in jd_map:
            if jd not in jd_order:
                jd_order.append(jd)

        for jd in jd_order:
            emp_list = jd_map.get(jd, [])
            if not emp_list:
                continue

            # ── Baris separator jobdesk ───────────────────────────────────────
            ws.merge_cells(start_row=current_row, start_column=1,
                           end_row=current_row,   end_column=TOTAL_COLS - 1)
            sc(ws.cell(current_row, 1, f'  ▸  {jd}  ({len(emp_list)} orang)'),
               bg=COLORS['jd_hdr'], bold=True, color='1A2A3A', h='left', size=9)
            ws.row_dimensions[current_row].height = 16
            current_row += 1

            # ── Baris pegawai ─────────────────────────────────────────────────
            for alt_idx, emp in enumerate(emp_list):
                write_emp_row(ws, current_row, emp, alt=alt_idx % 2 == 1)
                current_row += 1

        # Baris footer ringkasan
        current_row += 1
        total_emps = sum(len(v) for v in jd_map.values())
        footer = ws.cell(current_row, 1,
                         f'Total pegawai: {total_emps}  |  Total hari periode: {n_dates}')
        footer.font      = Font(bold=True, italic=True, size=9, color='555555')
        footer.alignment = Alignment(horizontal='left')

    # Fallback jika tidak ada cabang terdefinisi
    if not wb.worksheets:
        ws = wb.create_sheet(title='Jadwal')
        set_col_widths(ws)
        write_col_headers(ws, 1)
        for ri, emp in enumerate(emps, 2):
            write_emp_row(ws, ri, emp)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f'jadwal_{pk}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/schedule/<pk>/export/pdf')
def export_pdf(pk):
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.lib import colors as rc
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, PageBreak)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm

    period   = load_schedule(pk)
    emps     = load_employees()
    settings = load_settings()
    dates    = period['dates']
    sched    = period.get('schedule', {})
    branches = settings.get('branches', [])
    
    bid_filter = request.args.get('bid')
    if bid_filter:
        branches = [b for b in branches if b['id'] == bid_filter]
        
    day_abbr = ['Sen', 'Sel', 'Rab', 'Kam', 'Jum', 'Sab', 'Min']

    STATUS_COLORS = {
        'PAGI':  rc.HexColor('#FFD700'),
        'SIANG': rc.HexColor('#3A7BD5'),
        'OFF':   rc.HexColor('#E74C3C'),
        'CUTI':  rc.HexColor('#2ECC71'),
    }
    STATUS_SHORT = {'PAGI': 'P', 'SIANG': 'S', 'OFF': 'OFF', 'CUTI': 'CT'}

    # Pre-group
    by_bjd = {}
    for b in branches:
        by_bjd[b['id']] = {}
    for emp in emps:
        bid = emp.get('branch', '')
        jd  = emp['jobdesk']
        if bid not in by_bjd:
            by_bjd[bid] = {}
        by_bjd[bid].setdefault(jd, []).append(emp)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A3),
                            leftMargin=1*cm, rightMargin=1*cm,
                            topMargin=1.5*cm, bottomMargin=1*cm)

    title_st  = ParagraphStyle('T',  fontSize=13, spaceAfter=6,  alignment=1,
                                fontName='Helvetica-Bold')
    jd_st     = ParagraphStyle('JD', fontSize=9,  spaceAfter=2,
                                textColor=rc.HexColor('#1A2A3A'),
                                fontName='Helvetica-Bold')

    date_header = [
        f"{date.fromisoformat(d).strftime('%d/%m')}\n{day_abbr[date.fromisoformat(d).weekday()]}"
        for d in dates
    ]

    elements = []

    for bi, branch in enumerate(branches):
        bid    = branch['id']
        jd_map = by_bjd.get(bid, {})
        if not any(jd_map.values()):
            continue

        if bi > 0:
            elements.append(PageBreak())

        elements.append(Paragraph(
            f"JADWAL SHIFT  ·  {branch['name'].upper()}  ·  {period.get('label', pk).upper()}",
            title_st,
        ))

        jd_order = [jd for jd in branch.get('jobdesks', []) if jd in jd_map]
        for jd in jd_map:
            if jd not in jd_order:
                jd_order.append(jd)

        for jd in jd_order:
            emp_list = jd_map.get(jd, [])
            if not emp_list:
                continue

            elements.append(Paragraph(f'▸  {jd}  ({len(emp_list)} orang)', jd_st))

            header     = ['Nama', 'G'] + date_header + ['P', 'S', 'OFF', 'CT']
            table_data = [header]
            cstyles    = []

            for ri, emp in enumerate(emp_list, 1):
                eid    = emp['id']
                row    = [emp['name'], emp['gender']]
                counts = {'PAGI': 0, 'SIANG': 0, 'OFF': 0, 'CUTI': 0}
                for ci, d in enumerate(dates):
                    st  = sched.get(eid, {}).get(d)
                    row.append(STATUS_SHORT.get(st, '-'))
                    if st in STATUS_COLORS:
                        col = ci + 2
                        cstyles.append(('BACKGROUND', (col, ri), (col, ri), STATUS_COLORS[st]))
                        if st in ('SIANG', 'OFF'):
                            cstyles.append(('TEXTCOLOR', (col, ri), (col, ri), rc.white))
                    if st in counts:
                        counts[st] += 1
                row += [counts['PAGI'], counts['SIANG'], counts['OFF'], counts['CUTI']]
                table_data.append(row)

            col_widths = [3.5*cm, 0.55*cm] + [0.68*cm]*len(dates) + [0.7*cm]*4
            t = Table(table_data, colWidths=col_widths, repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND',    (0, 0), (-1, 0), rc.HexColor('#2C3E50')),
                ('TEXTCOLOR',     (0, 0), (-1, 0), rc.white),
                ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE',      (0, 0), (-1, -1), 7),
                ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID',          (0, 0), (-1, -1), 0.3, rc.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1),
                 [rc.white, rc.HexColor('#F8F9FA')]),
                ('ALIGN',         (0, 1), (0, -1), 'LEFT'),
                ('FONTNAME',      (0, 1), (-1, -1), 'Helvetica'),
            ] + cstyles))

            elements.append(t)
            elements.append(Spacer(1, 0.35*cm))

    if not elements:
        elements.append(Paragraph('Tidak ada data jadwal.', getSampleStyleSheet()['Normal']))

    doc.build(elements)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f'jadwal_{pk}.pdf',
        mimetype='application/pdf',
    )

# ─── Entry point ──────────────────────────────────────────────────────────────

if __name__ == '__main__':
    init_data()
    app.run(debug=True, port=5000)