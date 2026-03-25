from flask import render_template, request, jsonify, redirect, url_for, send_file
import os
import io
from datetime import date
from app.Models.DataStore import (
    load_schedule, save_schedule, load_settings, load_employees, 
    load_employees_for_period, group_by_branch, delete_schedule
)
from scheduler import get_period_dates, calculate_summary, generate_schedule

MONTH_NAMES = [
    '', 'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
    'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember'
]

class ScheduleController:
    @staticmethod
    def store():
        year   = int(request.form['year'])
        month  = int(request.form['month'])
        bid    = request.form.get('branch', '')
        pk = f'{year}-{month:02d}-{bid}' if bid else f'{year}-{month:02d}'

        period = load_schedule(pk)
        if not period:
            settings = load_settings()
            emps_all = load_employees()
            emps = [e for e in emps_all if e.get('branch') == bid] if bid else emps_all
            dates = get_period_dates(year, month)
            period = {
                'period_key': pk,
                'year':       year,
                'month':      month,
                'branch_id':  bid,
                'label':      f"{MONTH_NAMES[month]} {year}",
                'dates':      dates,
                'off_days':   {e['id']: [] for e in emps},
                'cuti_days':  {e['id']: [] for e in emps},
                'schedule':   {e['id']: {} for e in emps},
                'generated':  False,
            }
            save_schedule(pk, period)

        if period.get('generated'):
            return redirect(url_for('schedule_view', pk=pk))
        return redirect(url_for('schedule_setup', pk=pk))

    @staticmethod
    def setup(pk):
        period = load_schedule(pk)
        if not period:
            return redirect(url_for('index'))

        if period.get('generated'):
            return redirect(url_for('schedule_view', pk=pk))

        emps      = load_employees_for_period(period)
        settings  = load_settings()
        by_branch = group_by_branch(emps, settings)

        summary = calculate_summary(period, emps)
        return render_template('schedule_setup.html', period=period, by_branch=by_branch, summary=summary)

    @staticmethod
    def save_off(pk):
        period = load_schedule(pk)
        if not period:
            return jsonify({'success': False})
            
        eid = request.form.get('eid')
        if not eid:
            return jsonify({'success': False, 'error': 'No EID provided'})

        if 'off_days' not in period: period['off_days'] = {}
        if 'cuti_days' not in period: period['cuti_days'] = {}
        
        period['off_days'][eid]  = request.form.getlist('off[]')
        period['cuti_days'][eid] = request.form.getlist('cuti[]')
        
        if 'schedule' not in period: period['schedule'] = {}
        if eid not in period['schedule']: period['schedule'][eid] = {}
        
        for d in list(period['schedule'][eid].keys()):
            if period['schedule'][eid][d] in ('OFF', 'CUTI'):
                period['schedule'][eid].pop(d, None)
                
        for d in period['off_days'][eid]:
            period['schedule'][eid][d] = 'OFF'
        for d in period['cuti_days'][eid]:
            period['schedule'][eid][d] = 'CUTI'

        save_schedule(pk, period)
        return jsonify({'success': True})

    @staticmethod
    def auto_off_template(pk):
        period = load_schedule(pk)
        if not period:
            return jsonify({'success': False})
            
        emps = load_employees_for_period(period)
        settings = load_settings()
        by_branch = group_by_branch(emps, settings)
        
        if 'off_days' not in period: period['off_days'] = {e['id']: [] for e in emps}
        if 'schedule' not in period: period['schedule'] = {e['id']: {} for e in emps}
        
        for e in emps:
            period['off_days'][e['id']] = []
            for d in list(period['schedule'].get(e['id'], {}).keys()):
                if period['schedule'][e['id']][d] == 'OFF':
                    period['schedule'][e['id']].pop(d, None)
                    
        dates = period['dates']
        
        for b_id, b_data in by_branch.items():
            for jd, jd_emps in b_data['by_jobdesk'].items():
                jd_emps = sorted(jd_emps, key=lambda x: x['name'])
                for i, emp in enumerate(jd_emps):
                    eid = emp['id']
                    offset = i % 7
                    off_count = 0
                    for day_idx in range(offset, len(dates), 7):
                        if off_count >= 4:
                            break
                        d_str = dates[day_idx]
                        period['off_days'][eid].append(d_str)
                        period['schedule'][eid][d_str] = 'OFF'
                        off_count += 1

        save_schedule(pk, period)
        return jsonify({'success': True, 'reload': True, 'off_days': period['off_days']})

    @staticmethod
    def generate(pk):
        period = load_schedule(pk)
        if not period:
            return jsonify({'success': False, 'error': 'Period not found'})
        emps = load_employees_for_period(period)
        if not emps:
            return jsonify({'success': False, 'error': 'Tidak ada pegawai di cabang ini'})
            
        # Jika regenerate, hapus semua shift PAGI dan SIANG yang ada sebelumnya 
        # (OFF dan CUTI tetap akan dipertahankan oleh algoritma scheduler)
        if period.get('generated', False):
            if 'schedule' in period:
                for eid, shifts in period['schedule'].items():
                    for d in list(shifts.keys()):
                        if shifts[d] in ('PAGI', 'SIANG'):
                            del shifts[d]
                            
        sched = generate_schedule(period, emps)
        period['schedule']  = sched
        period['generated'] = True
        save_schedule(pk, period)
        return jsonify({'success': True, 'redirect': url_for('schedule_view', pk=pk)})

    @staticmethod
    def show(pk):
        period = load_schedule(pk)
        if not period:
            return redirect(url_for('index'))
        emps      = load_employees_for_period(period)
        settings  = load_settings()
        summary   = calculate_summary(period, emps)
        by_branch = group_by_branch(emps, settings)

        by_jd = {}
        for e in emps:
            by_jd.setdefault(e['jobdesk'], []).append(e)

        return render_template(
            'schedule_view.html',
            period=period, employees=emps,
            by_branch=by_branch,
            by_jobdesk=by_jd,
            settings=settings,
            summary=summary,
        )

    @staticmethod
    def edit_cell(pk):
        data   = request.get_json()
        period = load_schedule(pk)
        if not period:
            return jsonify({'success': False})
            
        eid    = data['emp_id']
        d      = data['date']
        status = data['status']
        
        if status == '':
            period['schedule'].setdefault(eid, {}).pop(d, None)
        else:
            period['schedule'].setdefault(eid, {})[d] = status
            
        off_list = period.setdefault('off_days', {}).setdefault(eid, [])
        cuti_list = period.setdefault('cuti_days', {}).setdefault(eid, [])
        
        if status == 'OFF':
            if d not in off_list: off_list.append(d)
            if d in cuti_list: cuti_list.remove(d)
        elif status == 'CUTI':
            if d not in cuti_list: cuti_list.append(d)
            if d in off_list: off_list.remove(d)
        else:
            if d in off_list: off_list.remove(d)
            if d in cuti_list: cuti_list.remove(d)
            
        save_schedule(pk, period)
        return jsonify({'success': True})

    @staticmethod
    def destroy(pk):
        delete_schedule(pk)
        return redirect(url_for('index'))

    @staticmethod
    def export_excel(pk):
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        period   = load_schedule(pk)
        emps     = load_employees()
        settings = load_settings()
        dates    = period['dates']
        sched    = period.get('schedule', {})
        branches = settings.get('branches', [])
        
        bid_filter = request.args.get('bid')
        if bid_filter:
            branches = [b for b in branches if b['id'] == bid_filter]
            
        n_dates  = len(dates)
        DATE_START_COL = 4
        SUM_START_COL  = DATE_START_COL + n_dates
        TOTAL_COLS     = SUM_START_COL + 4

        COLORS = {
            'PAGI':    'FFD700',
            'SIANG':   '3A7BD5',
            'OFF':     'E74C3C',
            'CUTI':    '2ECC71',
            'branch':  '1A252F',
            'col_hdr': '2C3E50',
            'jd_hdr':  'D5E8F7',
            'alt':     'F7FAFD',
        }
        day_abbr = ['Sen', 'Sel', 'Rab', 'Kam', 'Jum', 'Sab', 'Min']
        thin        = Side(border_style='thin', color='CCCCCC')
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        by_bjd = {}
        for b in branches:
            by_bjd[b['id']] = {}
        for emp in emps:
            bid = emp.get('branch', '')
            jd  = emp['jobdesk']
            if bid not in by_bjd:
                by_bjd[bid] = {}
            by_bjd[bid].setdefault(jd, []).append(emp)

        def sc(cell, bg=None, bold=False, color='000000', size=9, h='center', v='center', wrap=False):
            if bg:
                cell.fill = PatternFill('solid', fgColor=bg)
            cell.font      = Font(bold=bold, color=color, size=size)
            cell.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
            cell.border    = thin_border
            return cell

        def write_col_headers(ws, row):
            sc(ws.cell(row, 1, 'Nama'),    bg=COLORS['col_hdr'], bold=True, color='FFFFFF')
            sc(ws.cell(row, 2, 'Jobdesk'), bg=COLORS['col_hdr'], bold=True, color='FFFFFF')
            sc(ws.cell(row, 3, 'G'),       bg=COLORS['col_hdr'], bold=True, color='FFFFFF')
            for ci, d in enumerate(dates, DATE_START_COL):
                do = date.fromisoformat(d)
                sc(ws.cell(row, ci, f"{do.strftime('%d/%m')}\n{day_abbr[do.weekday()]}"),
                   bg=COLORS['col_hdr'], bold=True, color='FFFFFF', wrap=True)
            for i, lbl in enumerate(['PAGI', 'SIANG', 'OFF', 'CUTI']):
                sc(ws.cell(row, SUM_START_COL + i, lbl),
                   bg=COLORS['col_hdr'], bold=True, color='FFFFFF')
            ws.row_dimensions[row].height = 34

        def write_emp_row(ws, row, emp, alt=False):
            eid = emp['id']
            bg  = COLORS['alt'] if alt else None
            sc(ws.cell(row, 1, emp['name']),    bg=bg, h='left',   size=9)
            sc(ws.cell(row, 2, emp['jobdesk']), bg=bg, h='left',   size=9)
            sc(ws.cell(row, 3, emp['gender']),  bg=bg,             size=9)
            counts = {'PAGI': 0, 'SIANG': 0, 'OFF': 0, 'CUTI': 0}
            for ci, d in enumerate(dates, DATE_START_COL):
                st = sched.get(eid, {}).get(d, '')
                c  = sc(ws.cell(row, ci, st or '-'), bg=bg, size=8)
                if st in COLORS:
                    c.fill = PatternFill('solid', fgColor=COLORS[st])
                    txt    = 'FFFFFF' if st in ('SIANG', 'OFF') else '333333'
                    c.font = Font(size=8, bold=True, color=txt)
                    c.border = thin_border
                if st in counts:
                    counts[st] += 1
            for i, lbl in enumerate(['PAGI', 'SIANG', 'OFF', 'CUTI']):
                sc(ws.cell(row, SUM_START_COL + i, counts[lbl]), bg=bg, bold=True, size=9)
            ws.row_dimensions[row].height = 18

        def set_col_widths(ws):
            ws.column_dimensions['A'].width = 22
            ws.column_dimensions['B'].width = 13
            ws.column_dimensions['C'].width = 5
            for ci in range(DATE_START_COL, SUM_START_COL):
                ws.column_dimensions[get_column_letter(ci)].width = 6.5
            for ci in range(SUM_START_COL, SUM_START_COL + 4):
                ws.column_dimensions[get_column_letter(ci)].width = 7

        wb = Workbook()
        wb.remove(wb.active)

        for branch in branches:
            bid    = branch['id']
            jd_map = by_bjd.get(bid, {})
            if not any(jd_map.values()):
                continue

            sheet_name = branch['name'][:31].replace('/', '-').replace('\\', '-')
            ws = wb.create_sheet(title=sheet_name)
            set_col_widths(ws)

            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COLS - 1)
            title_text = f"JADWAL SHIFT  ·  {branch['name'].upper()}  ·  {period.get('label', pk).upper()}"
            sc(ws.cell(1, 1, title_text), bg=COLORS['branch'], bold=True, color='FFFFFF', size=13, wrap=False)
            ws.row_dimensions[1].height = 28

            write_col_headers(ws, 2)
            current_row = 3

            jd_order = [jd for jd in branch.get('jobdesks', []) if jd in jd_map]
            for jd in jd_map:
                if jd not in jd_order:
                    jd_order.append(jd)

            for jd in jd_order:
                emp_list = jd_map.get(jd, [])
                if not emp_list:
                    continue

                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=TOTAL_COLS - 1)
                sc(ws.cell(current_row, 1, f'  ▸  {jd}  ({len(emp_list)} orang)'), bg=COLORS['jd_hdr'], bold=True, color='1A2A3A', h='left', size=9)
                ws.row_dimensions[current_row].height = 16
                current_row += 1

                for alt_idx, emp in enumerate(emp_list):
                    write_emp_row(ws, current_row, emp, alt=alt_idx % 2 == 1)
                    current_row += 1

            current_row += 1
            total_emps = sum(len(v) for v in jd_map.values())
            footer = ws.cell(current_row, 1, f'Total pegawai: {total_emps}  |  Total hari periode: {n_dates}')
            footer.font      = Font(bold=True, italic=True, size=9, color='555555')
            footer.alignment = Alignment(horizontal='left')

        if not wb.worksheets:
            ws = wb.create_sheet(title='Jadwal')
            set_col_widths(ws)
            write_col_headers(ws, 1)
            for ri, emp in enumerate(emps, 2):
                write_emp_row(ws, ri, emp)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            as_attachment=True,
            download_name=f'jadwal_{pk}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    @staticmethod
    def export_pdf(pk):
        from reportlab.lib.pagesizes import A3, landscape
        from reportlab.lib import colors as rc
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm

        period   = load_schedule(pk)
        emps     = load_employees()
        settings = load_settings()
        dates    = period['dates']
        sched    = period.get('schedule', {})
        branches = settings.get('branches', [])
        
        bid_filter = request.args.get('bid')
        if bid_filter:
            branches = [b for b in branches if b['id'] == bid_filter]
            
        day_abbr = ['Sen', 'Sel', 'Rab', 'Kam', 'Jum', 'Sab', 'Min']

        STATUS_COLORS = {
            'PAGI':  rc.HexColor('#FFD700'),
            'SIANG': rc.HexColor('#3A7BD5'),
            'OFF':   rc.HexColor('#E74C3C'),
            'CUTI':  rc.HexColor('#2ECC71'),
        }
        STATUS_SHORT = {'PAGI': 'P', 'SIANG': 'S', 'OFF': 'OFF', 'CUTI': 'CT'}

        by_bjd = {}
        for b in branches:
            by_bjd[b['id']] = {}
        for emp in emps:
            bid = emp.get('branch', '')
            jd  = emp['jobdesk']
            if bid not in by_bjd:
                by_bjd[bid] = {}
            by_bjd[bid].setdefault(jd, []).append(emp)

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A3),
                                leftMargin=1*cm, rightMargin=1*cm,
                                topMargin=1.5*cm, bottomMargin=1*cm)

        title_st  = ParagraphStyle('T',  fontSize=13, spaceAfter=6,  alignment=1, fontName='Helvetica-Bold')
        jd_st     = ParagraphStyle('JD', fontSize=9,  spaceAfter=2, textColor=rc.HexColor('#1A2A3A'), fontName='Helvetica-Bold')

        date_header = [f"{date.fromisoformat(d).strftime('%d/%m')}\n{day_abbr[date.fromisoformat(d).weekday()]}" for d in dates]

        elements = []

        for bi, branch in enumerate(branches):
            bid    = branch['id']
            jd_map = by_bjd.get(bid, {})
            if not any(jd_map.values()):
                continue

            if bi > 0:
                elements.append(PageBreak())

            elements.append(Paragraph(f"JADWAL SHIFT  ·  {branch['name'].upper()}  ·  {period.get('label', pk).upper()}", title_st))

            jd_order = [jd for jd in branch.get('jobdesks', []) if jd in jd_map]
            for jd in jd_map:
                if jd not in jd_order:
                    jd_order.append(jd)

            for jd in jd_order:
                emp_list = jd_map.get(jd, [])
                if not emp_list:
                    continue

                elements.append(Paragraph(f'▸  {jd}  ({len(emp_list)} orang)', jd_st))

                header     = ['Nama', 'G'] + date_header + ['P', 'S', 'OFF', 'CT']
                table_data = [header]
                cstyles    = []

                for ri, emp in enumerate(emp_list, 1):
                    eid    = emp['id']
                    row    = [emp['name'], emp['gender']]
                    counts = {'PAGI': 0, 'SIANG': 0, 'OFF': 0, 'CUTI': 0}
                    for ci, d in enumerate(dates):
                        st  = sched.get(eid, {}).get(d)
                        row.append(STATUS_SHORT.get(st, '-'))
                        if st in STATUS_COLORS:
                            col = ci + 2
                            cstyles.append(('BACKGROUND', (col, ri), (col, ri), STATUS_COLORS[st]))
                            if st in ('SIANG', 'OFF'):
                                cstyles.append(('TEXTCOLOR', (col, ri), (col, ri), rc.white))
                        if st in counts:
                            counts[st] += 1
                    row += [counts['PAGI'], counts['SIANG'], counts['OFF'], counts['CUTI']]
                    table_data.append(row)

                col_widths = [3.5*cm, 0.55*cm] + [0.68*cm]*len(dates) + [0.7*cm]*4
                t = Table(table_data, colWidths=col_widths, repeatRows=1)
                t.setStyle(TableStyle([
                    ('BACKGROUND',    (0, 0), (-1, 0), rc.HexColor('#2C3E50')),
                    ('TEXTCOLOR',     (0, 0), (-1, 0), rc.white),
                    ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE',      (0, 0), (-1, -1), 7),
                    ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
                    ('GRID',          (0, 0), (-1, -1), 0.3, rc.grey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [rc.white, rc.HexColor('#F8F9FA')]),
                    ('ALIGN',         (0, 1), (0, -1), 'LEFT'),
                    ('FONTNAME',      (0, 1), (-1, -1), 'Helvetica'),
                ] + cstyles))

                elements.append(t)
                elements.append(Spacer(1, 0.35*cm))

        if not elements:
            elements.append(Paragraph('Tidak ada data jadwal.', getSampleStyleSheet()['Normal']))

        doc.build(elements)
        buf.seek(0)
        return send_file(
            buf,
            as_attachment=True,
            download_name=f'jadwal_{pk}.pdf',
            mimetype='application/pdf',
        )
